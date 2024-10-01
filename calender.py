from telegram_bot_calendar import DetailedTelegramCalendar, LSTEP
from telebot import types
import telebot
import openpyxl
import pandas as pd
import time
import sqlite3
# Импортируем нужный объект из библиотеки
from docxtpl import DocxTemplate
import emoji

token = '7526701581:AAEj5YAQJ8B-_VaKAEaGjdtw4ckL_aA1u-A'
bot = telebot.TeleBot(token)



key = ""
key_memo = 0000-00-00
file_path = 'training_report.xlsx'  # Укажите путь к вашему файлу


doc = DocxTemplate("Форма СЗ на обучение по ОТ (ОТ обучение 2).docx")

oneDate = 0
twoDate = 0
context = {
    'name': '',
    'profession': '',
    'division': '',
    'data': '',
    'email': '',
    'name1': '',
    'profession1': '',
    'division1': '',
    'data1': '',
    'email1': '',
    'name2': '',
    'profession2': '',
    'division2': '',
    'data2': '',
    'email2': '',
    'name3': '',
    'profession3': '',
    'division3': '',
    'data3': '',
    'email3': '',
    'name4': '',
    'profession4': '',
    'division4': '',
    'data4': '',
    'email4': ''
}

# SELECT MAX(date) AS max_date FROM training_report_OT

profession = 0
moth = -1
#

test_quition = telebot.types.ReplyKeyboardMarkup(True, True)
test_quition.row('1', '2', '3', '4', '5', '6')
test_quition.row('7', '8', '9', '10', '11', '12')

type_lerning = telebot.types.ReplyKeyboardMarkup(True, True)
type_lerning.row("ОТ 1", "ОТ 2", "ПБ (1-7)")

type_lerning_two = telebot.types.ReplyKeyboardMarkup(True, True)
type_lerning_two.row("ОТ", "ПБ")
# test_quition.row('7', '8', '9', '10', '11', '12')
# remove_keyboard = telebot.types.ReplyKeyboardRemove()

function0 = telebot.types.InlineKeyboardMarkup()
function1 = telebot.types.InlineKeyboardButton("Составить отчет по работнику", callback_data="employ")
function2 = telebot.types.InlineKeyboardButton("Составить отчет на год", callback_data="year")
function3 = telebot.types.InlineKeyboardButton("Добавить новое обучение", callback_data="calendar")
function4 = telebot.types.InlineKeyboardButton("Сформировать служебную записку (ОТ2)", callback_data="memo")
function5 = telebot.types.InlineKeyboardButton("Скоректировать обучение по срокам", callback_data="time")
function6 = telebot.types.InlineKeyboardButton("Новый работник", callback_data="newEmployee")
function7 = telebot.types.InlineKeyboardButton("Добавить информацию по обучению", callback_data="newLerning")
function8 = telebot.types.InlineKeyboardButton("Добавить работника на обучение", callback_data="employeeLerning")
function9 = telebot.types.InlineKeyboardButton("Cформировать отчет по своему подразделению за периуд времени", callback_data="document")
function0.add(function1).add(function2).add(function3).add(function4).add(function5).add(function7).add(function8).add(function9)



def clean_xlsx():
    workbook = openpyxl.load_workbook(file_path)
    # Перебираем все листы в книге
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]

        # Очищаем содержимое всех ячеек
        for row in worksheet.iter_rows():
            for cell in row:
                cell.value = None  # Устанавливаем значение ячейки в None

    # Сохраняем изменения в новый файл или перезаписываем существующий
    workbook.save('training_report.xlsx')  # Укажите имя для сохранения
    # workbook.save(file_path)  # Чтобы перезаписать оригинальный файл

    print("Документ очищен.")










@bot.message_handler(commands=['start'])
def start_message(message):
    # con = sqlite3.connect('CityDiscoveries.db')
    # cursorObj = con.cursor()
    # albums = (message.chat.id, 0)
    # cursorObj.execute("INSERT INTO men VALUES (?,?)", albums)
    # con.commit()
    print(message.chat.id)
    print(message.chat)
    bot.send_message(message.chat.id, 'Введите ваше ФИО')
    bot.send_message(message.chat.id, 'Вы авторизовались, как работник ОЦ')

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    button = types.KeyboardButton("Поделиться номером", request_contact=True)
    markup.add(button)
    bot.send_message(message.chat.id, "Привет! Нажми на кнопку ниже, чтобы поделиться своим номером:",
                     reply_markup=markup)


# Обработка получения контакта
@bot.message_handler(content_types=['contact'])
def handle_contact(message):
    global profession
    phone_number = message.contact.phone_number

    con = sqlite3.connect('BdTrainingCenter.db')
    cursor = con.cursor()
    # bot.send_message(message.chat.id, 'Введите ФИО сотрудника по которому не обходимо составить отчет')

    sqlite_select = f"""SELECT * FROM manager WHERE phone = {phone_number}"""
    cursor.execute(sqlite_select)

    if len(cursor.fetchall()) == 1 :
        profession = 1
        sqlite_select = f"""UPDATE manager SET tg = {message.chat.id} WHERE phone = {phone_number}"""
        cursor.execute(sqlite_select)
        con.commit()
        bot.send_message(message.chat.id, 'Успешная авторизация')

    else :
        bot.send_message(message.chat.id, 'Вас нет в списке сотрудников')
    # bot.send_message(message.chat.id, f"Спасибо! Ваш номер телефона: {phone_number}")



@bot.message_handler(commands=['menu'])
def menu(message):
    global profession
    if (profession == 0):
        print(message.chat.id)
        sqlite_select = f"""SELECT * FROM manager WHERE tg = {message.chat.id}"""
        con = sqlite3.connect('BdTrainingCenter.db')
        cursor = con.cursor()
        cursor.execute(sqlite_select)
        if (len(cursor.fetchall()) == 1):

            profession = 1

    else:
        menu(message.chat.id)

def menu(id):
    bot.send_message(id, text='Вот что я могу:', reply_markup=function0)
    global key_memo
    key_memo = 1
    # bot.send_message(id,
    #                  text='Если в списке нет интересующего вопроса, вы можете связаться с нами по телефону '
    #                       '+7 (800) 300-61-22')
    # random_fact(id)




def memoOT2(id, date):
    sqlite_select_tg = f"""SELECT * FROM manager WHERE tg = {id}"""
    con_tg = sqlite3.connect('BdTrainingCenter.db')
    cursor_tg = con_tg.cursor()
    cursor_tg.execute(sqlite_select_tg)
    tg = cursor_tg.fetchall()[0][2]
    print(tg)

    num = 0

    con = sqlite3.connect('BdTrainingCenter.db')
    cursor = con.cursor()

    sqlite_select = f"""SELECT employee.full_name, code_profession.Profession, employee.division, plane_OT.date, employee.email, plane_OT.number
FROM employee
JOIN plane_OT ON employee.ID = plane_OT.ID
JOIN code_profession ON employee.profession = code_profession.ID
WHERE employee.division = {tg} AND plane_OT.date = {date} AND plane_OT.number = 1
"""
    cursor.execute(sqlite_select)
    # for row_emp in cursorEmployee.fetchall():
    #     if row_emp[3] == 1:
    #         for row_emp in cursorEmployee.fetchall():

    # Данные для заполнения шаблона
    for row in cursor.fetchall():
        print(row)
        if num == 0:
            # 'name': '',
            #     'profession': '',
            #     'division': '',
            #     'data': '',
            #     'email': '',
            context['name'] = row[0]
            context['profession'] = row[1]
            print(context['profession'])
            context['division'] = row[2]
            context['data'] = row[3]
            context['email'] = row[4]
        elif num == 1:
            # 'name': '',
            #     'profession': '',
            #     'division': '',
            #     'data': '',
            #     'email': '',
            context['name1'] = row[0]
            context['profession1'] = row[1]
            context['division1'] = row[2]
            context['data1'] = row[3]
            context['email1'] = row[4]
        elif num == 2:
            # 'name': '',
            #     'profession': '',
            #     'division': '',
            #     'data': '',
            #     'email': '',
            context['name2'] = row[0]
            context['profession2'] = row[1]
            context['division2'] = row[2]
            context['data2'] = row[3]
            context['email2'] = row[4]
        elif num == 3:
            # 'name': '',
            #     'profession': '',
            #     'division': '',
            #     'data': '',
            #     'email': '',
            context['name3'] = row[0]
            context['profession3'] = row[1]
            context['division3'] = row[2]
            context['data3'] = row[3]
            context['email3'] = row[4]
        else:
            # 'name': '',
            #     'profession': '',
            #     'division': '',
            #     'data': '',
            #     'email': '',
            context['name4'] = row[0]
            context['profession4'] = row[1]
            context['division4'] = row[2]
            context['data4'] = row[3]
            context['email4'] = row[4]
            num = -1
            # Заполнение шаблона данными
            doc.render(context)

            # Сохранение документа
            doc.save("ОТ2.docx")
            with open('./ОТ2.docx', 'rb') as document:
                bot.send_document('5211807364', document)
            document.close()
            context['name'] = ''
            context['profession'] = ''
            print(context['profession'])
            context['division'] = ''
            context['data'] = ''
            context['email'] = ''

            context['name1'] = ''
            context['profession1'] = ''
            context['division1'] = ''
            context['data1'] = ''
            context['email1'] = ''

            context['name2'] = ''
            context['profession2'] = ''
            context['division2'] = ''
            context['data2'] = ''
            context['email2'] = ''

            context['name3'] = ''
            context['profession3'] = ''
            context['division3'] = ''
            context['data3'] = ''
            context['email3'] = ''

            context['name4'] = ''
            context['profession4'] = ''
            context['division4'] = ''
            context['data4'] = ''
            context['email4'] = ''

        num = num + 1

    if num != 0:
        # Заполнение шаблона данными
        doc.render(context)

        # Сохранение документа
        # doc.save("ОТ2"+ f"{row[0]}"+ ".docx")
        doc.save("ОТ2.docx")
        with open('./ОТ2.docx', 'rb') as document:
            bot.send_document('5211807364', document)
        document.close()
        context['name'] = ''
        context['profession'] = ''
        print(context['profession'])
        context['division'] = ''
        context['data'] = ''
        context['email'] = ''

        context['name1'] = ''
        context['profession1'] = ''
        context['division1'] = ''
        context['data1'] = ''
        context['email1'] = ''

        context['name2'] = ''
        context['profession2'] = ''
        context['division2'] = ''
        context['data2'] = ''
        context['email2'] = ''

        context['name3'] = ''
        context['profession3'] = ''
        context['division3'] = ''
        context['data3'] = ''
        context['email3'] = ''

        context['name4'] = ''
        context['profession4'] = ''
        context['division4'] = ''
        context['data4'] = ''
        context['email4'] = ''

    # df.to_excel('./training_report.xlsx', sheet_name=message.text, index=False)
    # doc = open(r'D:\проекты питон\severalmaz_training_center_bot\training_report.xlsx',
    #                            'rb')
    # bot.send_document(message.from_user.id, doc)


def memoOT1(id, date):
    sqlite_select_tg = f"""SELECT * FROM manager WHERE tg = {id}"""
    con_tg = sqlite3.connect('BdTrainingCenter.db')
    cursor_tg = con_tg.cursor()
    cursor_tg.execute(sqlite_select_tg)
    tg = cursor_tg.fetchall()[0][2]
    full_name = cursor_tg.fetchall()[0][1]
    print(tg)

    num = 0

    con = sqlite3.connect('BdTrainingCenter.db')
    cursor = con.cursor()

    sqlite_select = f"""SELECT lerning_division.date, employee.full_name, code_profession.Profession, employee.division, employee.email
    FROM employee
    JOIN lerning_division ON lerning_division.month = plane_OT.date
    JOIN plane_OT ON employee.ID = plane_OT.ID
    JOIN code_profession ON employee.profession = code_profession.ID
    WHERE employee.division = {tg} AND plane_OT.date = {date}"""
    cursor.execute(sqlite_select)
    # for row_emp in cursorEmployee.fetchall():
    #     if row_emp[3] == 1:
    #         for row_emp in cursorEmployee.fetchall():

    # Данные для заполнения шаблона
    for row in cursor.fetchall():
        print(row)
        if num == 0:
            # 'name': '',
            #     'profession': '',
            #     'division': '',
            #     'data': '',
            #     'email': '',
            context['name'] = row[0]
            context['profession'] = row[1]
            print(context['profession'])
            context['division'] = row[2]
            context['data'] = row[3]
            context['email'] = row[4]
        elif num == 1:
            # 'name': '',
            #     'profession': '',
            #     'division': '',
            #     'data': '',
            #     'email': '',
            context['name1'] = row[0]
            context['profession1'] = row[1]
            context['division1'] = row[2]
            context['data1'] = row[3]
            context['email1'] = row[4]
        elif num == 2:
            # 'name': '',
            #     'profession': '',
            #     'division': '',
            #     'data': '',
            #     'email': '',
            context['name2'] = row[0]
            context['profession2'] = row[1]
            context['division2'] = row[2]
            context['data2'] = row[3]
            context['email2'] = row[4]
        elif num == 3:
            # 'name': '',
            #     'profession': '',
            #     'division': '',
            #     'data': '',
            #     'email': '',
            context['name3'] = row[0]
            context['profession3'] = row[1]
            context['division3'] = row[2]
            context['data3'] = row[3]
            context['email3'] = row[4]
        else:
            # 'name': '',
            #     'profession': '',
            #     'division': '',
            #     'data': '',
            #     'email': '',
            context['name4'] = row[0]
            context['profession4'] = row[1]
            context['division4'] = row[2]
            context['data4'] = row[3]
            context['email4'] = row[4]
            num = -1
            # Заполнение шаблона данными
            doc.render(context)

            # Сохранение документа
            doc.save("ОТ2.docx")
            with open('./ОТ2.docx', 'rb') as document:
                bot.send_document('5211807364', document)
            document.close()
            context['name'] = ''
            context['profession'] = ''
            print(context['profession'])
            context['division'] = ''
            context['data'] = ''
            context['email'] = ''

            context['name1'] = ''
            context['profession1'] = ''
            context['division1'] = ''
            context['data1'] = ''
            context['email1'] = ''

            context['name2'] = ''
            context['profession2'] = ''
            context['division2'] = ''
            context['data2'] = ''
            context['email2'] = ''

            context['name3'] = ''
            context['profession3'] = ''
            context['division3'] = ''
            context['data3'] = ''
            context['email3'] = ''

            context['name4'] = ''
            context['profession4'] = ''
            context['division4'] = ''
            context['data4'] = ''
            context['email4'] = ''

        num = num + 1

    if num != 0:
        # Заполнение шаблона данными
        doc.render(context)

        # Сохранение документа
        # doc.save("ОТ2"+ f"{row[0]}"+ ".docx")
        doc.save("ОТ2.docx")
        with open('./ОТ2.docx', 'rb') as document:
            bot.send_document('5211807364', document)
        document.close()
        context['name'] = ''
        context['profession'] = ''
        print(context['profession'])
        context['division'] = ''
        context['data'] = ''
        context['email'] = ''

        context['name1'] = ''
        context['profession1'] = ''
        context['division1'] = ''
        context['data1'] = ''
        context['email1'] = ''

        context['name2'] = ''
        context['profession2'] = ''
        context['division2'] = ''
        context['data2'] = ''
        context['email2'] = ''

        context['name3'] = ''
        context['profession3'] = ''
        context['division3'] = ''
        context['data3'] = ''
        context['email3'] = ''

        context['name4'] = ''
        context['profession4'] = ''
        context['division4'] = ''
        context['data4'] = ''
        context['email4'] = ''

    # df.to_excel('./training_report.xlsx', sheet_name=message.text, index=False)
    # doc = open(r'D:\проекты питон\severalmaz_training_center_bot\training_report.xlsx',
    #                            'rb')
    # bot.send_document(message.from_user.id, doc)




@bot.message_handler(commands=['date'])
def date(m):
    calendar, step = DetailedTelegramCalendar().build()
    bot.send_message(m.chat.id,
                     f"Select {LSTEP[step]}",
                     reply_markup=calendar)


@bot.callback_query_handler(func=DetailedTelegramCalendar.func())
def cal(c):
    global oneDate, twoDate
    result, key, step = DetailedTelegramCalendar().process(c.data)
    if not result and key:
        bot.edit_message_text(f"Select {LSTEP[step]}",
                              c.message.chat.id,
                              c.message.message_id,
                              reply_markup=key)
    elif result:
        bot.edit_message_text(f"You selected {result}",
                              c.message.chat.id,
                              c.message.message_id)
        if (oneDate == 0000-00-00):
            oneDate = result
            bot.edit_message_text("Выберите дату окончания периуда",
                                  c.message.chat.id,
                                  c.message.message_id)
            date(c.message)
        else:
            twoDate = result
            sqlite_select = f"""SELECT * FROM manager WHERE tg = {c.message.chat.id}"""
            con = sqlite3.connect('BdTrainingCenter.db')
            cursor = con.cursor()
            cursor.execute(sqlite_select)
            division = 0
            for row in cursor:
                division = row[2]


#             Заполнить отчет данными о прошедших и будущих обучения за указанный периуд
# SELECT employee.full_name,code_profession.profession, training_report_OT.curse_profession, training_report_OT.curse_num, training_report_OT.date
#     FROM employee
#     JOIN training_report_OT ON employee.ID = training_report_OT.ID_employee
#     JOIN code_profession ON code_profession.ID = training_report_OT.curse_profession
#     WHERE employee.division = 1 AND training_report_OT.date > '2022-03-30' AND training_report_OT.date < '2025-03-30'

# Отчет по ОТ прошедших


            df = pd.DataFrame(columns=['Сотрудники',
                                        'Номер обучения ОТ',
                                       'Должность на которую проходили обучение',
                                       'Реквизиты документа',
                                       'Дата'])

            con = sqlite3.connect('BdTrainingCenter.db')

            cursorProfession = con.cursor()
            sqlite_select_Profession = f"""SELECT employee.full_name,training_report_OT.curse_num, code_profession.profession, training_report_OT.details_document, training_report_OT.date
    FROM employee
    JOIN training_report_OT ON employee.ID = training_report_OT.ID_employee 
    JOIN code_profession ON code_profession.ID = training_report_OT.curse_profession
    WHERE employee.division = {division} AND training_report_OT.date > '2022-03-30' AND training_report_OT.date < '2025-03-30'"""
            cursorProfession.execute(sqlite_select_Profession)
            for row_pro in cursorProfession.fetchall():
                new_row = pd.DataFrame({
                                'Сотрудники': row_pro[0],
                                'Номер обучения ОТ': [row_pro[1]],
                                'Должность на которую проходили обучение': [row_pro[2]],
                                'Реквизиты документа' : [row_pro[3]],
                                'Дата': [row_pro[4]]
                            }, index=[0])

                # Используем pd.concat для добавления новой строки
                df = pd.concat([df, new_row], ignore_index=True)

                # print(df)

            cursorProfession = con.cursor()
            sqlite_select_Profession = f"""SELECT employee.full_name,plane_OT.number, code_profession.profession, lerning_division.date
                FROM employee
                JOIN plane_OT ON employee.ID = plane_OT.ID
                JOIN lerning_division ON lerning_division.month = plane_OT.date
                JOIN code_profession ON code_profession.ID = plane_OT.profession
                WHERE employee.division = {division} AND lerning_division.start > '{oneDate}' AND lerning_division.start < '{twoDate}'"""
            cursorProfession.execute(sqlite_select_Profession)
            for row_pro in cursorProfession.fetchall():
                new_row = pd.DataFrame({
                    'Сотрудники': row_pro[0],
                    'Номер обучения ОТ': [row_pro[1]],
                    'Должность на которую проходили обучение': [row_pro[2]],
                    'Реквизиты документа': ["Не получен"],
                    'Дата': [row_pro[3]]
                }, index=[0])

                # Используем pd.concat для добавления новой строки
                df = pd.concat([df, new_row], ignore_index=True)

                # print(df)

            df.to_excel('./Отчет_ОТ_за_периуд.xlsx', sheet_name=c.message.text, index=False)
            doc = open(r'D:\проекты питон\severalmaz_training_center_bot\Отчет_ОТ_за_периуд.xlsx',
                       'rb')
            bot.send_document(c.message.chat.id, doc)
            doc.close()
            # Формирование для ПБ

            df = pd.DataFrame(columns=['Сотрудники',
                                       'Номер обучения ПБ',
                                       'Должность',
                                       'Реквизиты документа',
                                       'Дата'])

            con = sqlite3.connect('BdTrainingCenter.db')

            cursorProfession = con.cursor()
            sqlite_select_Profession = f"""SELECT employee.full_name,training_report_PB.curse_num, code_profession.profession, training_report_PB.details_document, training_report_PB.date
            FROM employee
            JOIN training_report_PB ON employee.ID = training_report_PB.ID_employee 
            JOIN code_profession ON code_profession.ID = employee.profession
            WHERE employee.division = {division} AND training_report_PB.date > '{oneDate}' AND training_report_PB.date < '{twoDate}'"""
            cursorProfession.execute(sqlite_select_Profession)
            for row_pro in cursorProfession.fetchall():
                new_row = pd.DataFrame({
                    'Сотрудники': row_pro[0],
                    'Номер обучения ПБ': [row_pro[1]],
                    'Должность': [row_pro[2]],
                    'Реквизиты документа': [row_pro[3]],
                    'Дата': [row_pro[4]]
                }, index=[0])

                # Используем pd.concat для добавления новой строки
                df = pd.concat([df, new_row], ignore_index=True)

                # print(df)

            cursorProfession = con.cursor()
            sqlite_select_Profession = f"""SELECT employee.full_name,plane_PB.number, code_profession.profession, lerning_division.date
                        FROM employee
                        JOIN plane_PB ON employee.ID = plane_PB.ID
                        JOIN lerning_division ON lerning_division.month = plane_PB.date
                        JOIN code_profession ON code_profession.ID = employee.profession
                        WHERE employee.division = {division} AND lerning_division.start > '{oneDate}' AND lerning_division.start < '{twoDate}'"""
            cursorProfession.execute(sqlite_select_Profession)
            for row_pro in cursorProfession.fetchall():
                new_row = pd.DataFrame({
                    'Сотрудники': row_pro[0],
                    'Номер обучения ПБ': [row_pro[1]],
                    'Должность': [row_pro[2]],
                    'Реквизиты документа': ["Не получен"],
                    'Дата': [row_pro[3]]
                }, index=[0])

                # Используем pd.concat для добавления новой строки
                df = pd.concat([df, new_row], ignore_index=True)

                # print(df)

            df.to_excel('./Отчет_ПБ_за_периуд.xlsx', sheet_name=c.message.text, index=False)
            doc = open(r'D:\проекты питон\severalmaz_training_center_bot\Отчет_ПБ_за_периуд.xlsx',
                       'rb')
            bot.send_document(c.message.chat.id, doc)
            doc.close()
            oneDate = 0000-00-00
            twoDate = 0000-00-00
            return



# Кнопки функций
@bot.callback_query_handler(func=lambda call: True)
def query_handler(call):
    global key
    if (call.data == "employ"):
        key = "employ"
        bot.send_message(call.message.chat.id, 'Введите ФИО работника')
    elif (call.data == "memo"):
        key = "memo"
        bot.send_message(call.message.chat.id, "Укажите номер месяца за, на который необходимо сформировать документы",
                         reply_markup=test_quition)

    elif (call.data == "employeeLerning"):
        key = "employeeLerning"
        bot.send_message(call.message.chat.id, 'Введите ФИО работника')
        date()
        # memo(call.message.chat.id)




    elif (call.data == "document"):
        bot.send_message(call.message.chat.id, 'Выберите дату начала периуда')
        date(call.message)
    #     key = "document"
    #     bot.send_message(call.message.chat.id, 'Напишите год, закоторый необходимо составить отчет')
#



@bot.message_handler(content_types=['text'])
def text(message):
    if (message.text == "ОТ 1"):
        memoOT2(message.chat.id, moth)
    elif key == "memo":
        if (message.text.isdigit()):
            if (int(message.text) >= 1 and int(message.text) < 13):
                # if (time.time())
                bot.send_message(message.chat.id, "По какому типу обучения составить записку?",
                         reply_markup=type_lerning)
    elif (key == "employ"):
        print(message.text)
        con = sqlite3.connect('BdTrainingCenter.db')
        cursorEmployee = con.cursor()
        bot.send_message(message.chat.id, 'Введите ФИО сотрудника по которому не обходимо составить отчет')
        sqlite_select_employee = """SELECT * from employee"""
        cursorEmployee.execute(sqlite_select_employee)
        for row_emp in cursorEmployee.fetchall():
            if row_emp[1] == message.text:
                id = row_emp[0]
                cursorReport = con.cursor()
                sqlite_select_report = """SELECT * from training_report"""
                cursorReport.execute(sqlite_select_report)

                # Открываем существующий файл
                # clean_xlsx()

                # df = pd.DataFrame({'Сотрудники': [],
                #                    'Должность организации': [],
                #                    'Реквизиты документа': [],
                #                    'Дата': []})
                df = pd.DataFrame(columns=['Сотрудники',
                                   'Должность организации',
                                   'Реквизиты документа',
                                   'Дата'])

                for row_rep in cursorReport.fetchall():
                    if row_rep[0] == id:
                        # print(row_rep)

                        cursorProfession = con.cursor()
                        sqlite_select_Profession = """SELECT * from code_profession"""
                        cursorProfession.execute(sqlite_select_Profession)
                        for row_pro in cursorProfession.fetchall():
                            if row_emp[2] == row_pro[0]:
                                # print(row_pro)
                                # df = pd.DataFrame(columns=['Имя', 'Возраст', 'Город'])

                                # Добавляем строки
                                # df = df.append({'Имя': 'Alice', 'Возраст': 25, 'Город': 'New York'}, ignore_index=True)
                                # df = pd.DataFrame({'Сотрудники': [message.text],
                                #                    'Должность организации': [row_pro[1]],
                                #                    'Реквизиты документа': [row_rep[1]],
                                #                    'Дата': row_rep[3]})
                                # df.loc= [[message.text], [row_pro[1]],[row_rep[1]],row_rep[3]]
                                # df = df.append({'Сотрудники': [message.text],
                                #                    'Должность организации': [row_pro[1]],
                                #                    'Реквизиты документа': [row_rep[1]],
                                #                    'Дата': row_rep[3]})

                                # df = df.append(pd.DataFrame({'Сотрудники': [message.text],
                                #                              'Должность организации': [row_pro[1]],
                                #                              'Реквизиты документа': [row_rep[1]],
                                #                              'Дата': [row_rep[3]]}), ignore_index=True)

                                # Создаем новый DataFrame для добавляемой строки
                                new_row = pd.DataFrame({
                                    'Сотрудники': [message.text],
                                    'Должность организации': [row_pro[1]],
                                    'Реквизиты документа': [row_rep[1]],
                                    'Дата': [row_rep[3]]
                                })

                                # Используем pd.concat для добавления новой строки
                                df = pd.concat([df, new_row], ignore_index=True)

                                # print(df)

                df.to_excel('./training_report.xlsx', sheet_name=message.text, index=False)
                doc = open(r'D:\проекты питон\severalmaz_training_center_bot\training_report.xlsx',
                           'rb')
                bot.send_document(message.from_user.id, doc)
                doc.close()

                return

    elif (key == "employeeLerning"):
        con1 = sqlite3.connect('BdTrainingCenter.db')
        cursor1 = con1.cursor()
        sqlite_select = f"""SELECT ID FROM employee WHERE full_name = {message.text} """
        cursor1.execute(sqlite_select)
        con1.commit()
        if len(cursor1.fetchall()) == 0:
            bot.send_message(message.chat.id, "Работника с данным ФИО нет, проверти правописание и повторите попытку")
        else:
            bot.send_message(message.chat.id, "Записать на какой тип обучения",
                             reply_markup=type_lerning_two)


    # elif (key == "document"):
    #     if (message.text.isdigit()):
    #         bot.send_message(message.chat.id, 'Выберите дату начала периуда')
    #         date(message)








if __name__ == "__main__":
    bot.polling(none_stop=True)