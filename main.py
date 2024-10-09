import telebot
import sqlite3
import openpyxl
import pandas as pd
from telebot import types
# Импортируем нужный объект из библиотеки
from docxtpl import DocxTemplate
import emoji

token = '7526701581:AAEj5YAQJ8B-_VaKAEaGjdtw4ckL_aA1u-A'
bot = telebot.TeleBot(token)
key = ""
key_memo = 0
file_path = 'training_report.xlsx'  # Укажите путь к вашему файлу


doc = DocxTemplate("Форма СЗ на обучение по ОТ (ОТ обучение 2).docx")


context = {
    'name': '',
    'profession': '',
    'division': '',
    'data': '',
    'email': '',
    'name1': '',
    'profession1': '',
    'division1': '',
    'data1': '',
    'email1': '',
    'name2': '',
    'profession2': '',
    'division2': '',
    'data2': '',
    'email2': '',
    'name3': '',
    'profession3': '',
    'division3': '',
    'data3': '',
    'email3': '',
    'name4': '',
    'profession4': '',
    'division4': '',
    'data4': '',
    'email4': ''
}

# SELECT MAX(date) AS max_date FROM training_report_OT

profession = 0
month = -1
#

test_quition = telebot.types.ReplyKeyboardMarkup(True, True)
test_quition.row('1', '2', '3', '4', '5', '6')
test_quition.row('7', '8', '9', '10', '11', '12')

type_lerning = telebot.types.ReplyKeyboardMarkup(True, True)
type_lerning.row("ОТ 1", "ОТ 2", "ПБ")
# test_quition.row('7', '8', '9', '10', '11', '12')
# remove_keyboard = telebot.types.ReplyKeyboardRemove()

function0 = telebot.types.InlineKeyboardMarkup()
function1 = telebot.types.InlineKeyboardButton("Составить отчет по работнику", callback_data="employ")
function2 = telebot.types.InlineKeyboardButton("Составить отчет за год", callback_data="year")
function3 = telebot.types.InlineKeyboardButton("Добавить новое обучение или скоректировать по срокам", callback_data="calendar")
function4 = telebot.types.InlineKeyboardButton("Сформировать служебную записку (ОТ2)", callback_data="memo")
# function5 = telebot.types.InlineKeyboardButton("Как купить тур?", callback_data="count")
# function6 = telebot.types.InlineKeyboardButton("Почему мне надо посетить город открытий?", callback_data="go")
# function7 = telebot.types.InlineKeyboardButton("Интересные факты о разных отраслях Москвы и почему её необходимо посетить", callback_data="moscow")
# function8 = telebot.types.InlineKeyboardButton("Тест на профориентацию", callback_data="profi")
# function9 = telebot.types.InlineKeyboardButton("Викторина", callback_data="victorina")
function0.add(function1).add(function2).add(function3).add(function4)#.add(function5).add(function7).add(function8).add(function9)



def clean_xlsx():
    workbook = openpyxl.load_workbook(file_path)

    # Перебираем все листы в книге
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]

        # Очищаем содержимое всех ячеек
        for row in worksheet.iter_rows():
            for cell in row:
                cell.value = None  # Устанавливаем значение ячейки в None

    # Сохраняем изменения в новый файл или перезаписываем существующий
    workbook.save('training_report.xlsx')  # Укажите имя для сохранения
    # workbook.save(file_path)  # Чтобы перезаписать оригинальный файл

    print("Документ очищен.")



@bot.message_handler(commands=['start'])
def start_message(message):
    # con = sqlite3.connect('CityDiscoveries.db')
    # cursorObj = con.cursor()
    # albums = (message.chat.id, 0)
    # cursorObj.execute("INSERT INTO men VALUES (?,?)", albums)
    # con.commit()
    print(message.chat.id)
    print(message.chat)
    bot.send_message(message.chat.id, 'Введите ваше ФИО')
    bot.send_message(message.chat.id, 'Вы авторизовались, как работник ОЦ')

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    button = types.KeyboardButton("Поделиться номером", request_contact=True)
    markup.add(button)
    bot.send_message(message.chat.id, "Привет! Нажми на кнопку ниже, чтобы поделиться своим номером:",
                     reply_markup=markup)



# Обработка получения контакта
@bot.message_handler(content_types=['contact'])
def handle_contact(message):
    global profession
    phone_number = message.contact.phone_number

    con = sqlite3.connect('BdTrainingCenter.db')
    cursor = con.cursor()
    # bot.send_message(message.chat.id, 'Введите ФИО сотрудника по которому не обходимо составить отчет')

    sqlite_select = f"""SELECT * FROM manager WHERE phone = {phone_number}"""
    cursor.execute(sqlite_select)

    if len(cursor.fetchall()) == 1 :
        sqlite_select = f"""UPDATE manager SET tg = {message.chat.id} WHERE phone = {phone_number}"""
        cursor.execute(sqlite_select)
        con.commit()
        bot.send_message(message.chat.id, 'Успешная авторизация')
        profession = 1

    else :
        bot.send_message(message.chat.id, 'Вас нет в списке сотрудников')
    # bot.send_message(message.chat.id, f"Спасибо! Ваш номер телефона: {phone_number}")



@bot.message_handler(commands=['menu'])
def menu(message):
    global profession
    if (profession == 0):
        print(message.chat.id)
        sqlite_select = f"""SELECT * FROM manager WHERE tg = {message.chat.id}"""
        con = sqlite3.connect('BdTrainingCenter.db')
        cursor = con.cursor()
        cursor.execute(sqlite_select)
        if (len(cursor.fetchall()) == 1):
            profession = 1

    if (profession == 1):
        menu(message.chat.id)




def memoOT2(id, date):
    sqlite_select_tg = f"""SELECT * FROM manager WHERE tg = {id}"""
    con_tg = sqlite3.connect('BdTrainingCenter.db')
    cursor_tg = con_tg.cursor()
    cursor_tg.execute(sqlite_select_tg)
    tg = cursor_tg.fetchall()[0][2]
    print(tg)
    print(date)

    num = 0

    con = sqlite3.connect('BdTrainingCenter.db')
    cursor = con.cursor()

    sqlite_select = f"""SELECT employee.full_name, code_profession.Profession, employee.division, plane_OT.date, employee.email, plane_OT.number
FROM employee
JOIN plane_OT ON employee.ID = plane_OT.ID
JOIN code_profession ON employee.profession = code_profession.ID
WHERE employee.division = {tg} AND plane_OT.date = {date} AND plane_OT.number = 1
"""
    cursor.execute(sqlite_select)
    # for row_emp in cursorEmployee.fetchall():
    #     if row_emp[3] == 1:
    #         for row_emp in cursorEmployee.fetchall():

    # Данные для заполнения шаблона
    for row in cursor.fetchall():
        print(row)
        if num == 0:
            # 'name': '',
            #     'profession': '',
            #     'division': '',
            #     'data': '',
            #     'email': '',
            context['name'] = row[0]
            context['profession'] = row[1]
            print(context['profession'])
            context['division'] = row[2]
            context['data'] = row[3]
            context['email'] = row[4]
        elif num == 1:
            # 'name': '',
            #     'profession': '',
            #     'division': '',
            #     'data': '',
            #     'email': '',
            context['name1'] = row[0]
            context['profession1'] = row[1]
            context['division1'] = row[2]
            context['data1'] = row[3]
            context['email1'] = row[4]
        elif num == 2:
            # 'name': '',
            #     'profession': '',
            #     'division': '',
            #     'data': '',
            #     'email': '',
            context['name2'] = row[0]
            context['profession2'] = row[1]
            context['division2'] = row[2]
            context['data2'] = row[3]
            context['email2'] = row[4]
        elif num == 3:
            # 'name': '',
            #     'profession': '',
            #     'division': '',
            #     'data': '',
            #     'email': '',
            context['name3'] = row[0]
            context['profession3'] = row[1]
            context['division3'] = row[2]
            context['data3'] = row[3]
            context['email3'] = row[4]
        else:
            # 'name': '',
            #     'profession': '',
            #     'division': '',
            #     'data': '',
            #     'email': '',
            context['name4'] = row[0]
            context['profession4'] = row[1]
            context['division4'] = row[2]
            context['data4'] = row[3]
            context['email4'] = row[4]
            num = -1
            # Заполнение шаблона данными
            doc.render(context)

            # Сохранение документа
            doc.save("ОТ2.docx")
            with open('./ОТ2.docx', 'rb') as document:
                bot.send_document('5211807364', document)
            document.close()
            context['name'] = ''
            context['profession'] = ''
            print(context['profession'])
            context['division'] = ''
            context['data'] = ''
            context['email'] = ''

            context['name1'] = ''
            context['profession1'] = ''
            context['division1'] = ''
            context['data1'] = ''
            context['email1'] = ''

            context['name2'] = ''
            context['profession2'] = ''
            context['division2'] = ''
            context['data2'] = ''
            context['email2'] = ''

            context['name3'] = ''
            context['profession3'] = ''
            context['division3'] = ''
            context['data3'] = ''
            context['email3'] = ''

            context['name4'] = ''
            context['profession4'] = ''
            context['division4'] = ''
            context['data4'] = ''
            context['email4'] = ''

        num = num + 1

    if num != 0:
        # Заполнение шаблона данными
        doc.render(context)

        # Сохранение документа
        # doc.save("ОТ2"+ f"{row[0]}"+ ".docx")
        doc.save("ОТ2.docx")
        with open('./ОТ2.docx', 'rb') as document:
            bot.send_document('5211807364', document)
        document.close()
        context['name'] = ''
        context['profession'] = ''
        print(context['profession'])
        context['division'] = ''
        context['data'] = ''
        context['email'] = ''

        context['name1'] = ''
        context['profession1'] = ''
        context['division1'] = ''
        context['data1'] = ''
        context['email1'] = ''

        context['name2'] = ''
        context['profession2'] = ''
        context['division2'] = ''
        context['data2'] = ''
        context['email2'] = ''

        context['name3'] = ''
        context['profession3'] = ''
        context['division3'] = ''
        context['data3'] = ''
        context['email3'] = ''

        context['name4'] = ''
        context['profession4'] = ''
        context['division4'] = ''
        context['data4'] = ''
        context['email4'] = ''

    # df.to_excel('./training_report.xlsx', sheet_name=message.text, index=False)
    # doc = open(r'D:\проекты питон\severalmaz_training_center_bot\training_report.xlsx',
    #                            'rb')
    # bot.send_document(message.from_user.id, doc)


def memoOT1(id, date):
    sqlite_select_tg = f"""SELECT * FROM manager WHERE tg = {id}"""
    con_tg = sqlite3.connect('BdTrainingCenter.db')
    cursor_tg = con_tg.cursor()
    cursor_tg.execute(sqlite_select_tg)
    tg = cursor_tg.fetchall()[0][2]
    print(tg)

    num = 0

    con = sqlite3.connect('BdTrainingCenter.db')
    cursor = con.cursor()

    sqlite_select = f"""SELECT employee.full_name, code_profession.Profession, employee.division, plane_OT.date, employee.email
    FROM employee
    JOIN plane_OT ON employee.ID = plane_OT.ID
    JOIN code_profession ON employee.profession = code_profession.ID
    WHERE employee.division = {tg} AND plane_OT.date = {date}"""
    cursor.execute(sqlite_select)
    # for row_emp in cursorEmployee.fetchall():
    #     if row_emp[3] == 1:
    #         for row_emp in cursorEmployee.fetchall():

    # Данные для заполнения шаблона
    for row in cursor.fetchall():
        print(row)
        if num == 0:
            # 'name': '',
            #     'profession': '',
            #     'division': '',
            #     'data': '',
            #     'email': '',
            context['name'] = row[0]
            context['profession'] = row[1]
            print(context['profession'])
            context['division'] = row[2]
            context['data'] = row[3]
            context['email'] = row[4]
        elif num == 1:
            # 'name': '',
            #     'profession': '',
            #     'division': '',
            #     'data': '',
            #     'email': '',
            context['name1'] = row[0]
            context['profession1'] = row[1]
            context['division1'] = row[2]
            context['data1'] = row[3]
            context['email1'] = row[4]
        elif num == 2:
            # 'name': '',
            #     'profession': '',
            #     'division': '',
            #     'data': '',
            #     'email': '',
            context['name2'] = row[0]
            context['profession2'] = row[1]
            context['division2'] = row[2]
            context['data2'] = row[3]
            context['email2'] = row[4]
        elif num == 3:
            # 'name': '',
            #     'profession': '',
            #     'division': '',
            #     'data': '',
            #     'email': '',
            context['name3'] = row[0]
            context['profession3'] = row[1]
            context['division3'] = row[2]
            context['data3'] = row[3]
            context['email3'] = row[4]
        else:
            # 'name': '',
            #     'profession': '',
            #     'division': '',
            #     'data': '',
            #     'email': '',
            context['name4'] = row[0]
            context['profession4'] = row[1]
            context['division4'] = row[2]
            context['data4'] = row[3]
            context['email4'] = row[4]
            num = -1
            # Заполнение шаблона данными
            doc.render(context)

            # Сохранение документа
            doc.save("ОТ2.docx")
            with open('./ОТ2.docx', 'rb') as document:
                bot.send_document('5211807364', document)
            document.close()
            context['name'] = ''
            context['profession'] = ''
            print(context['profession'])
            context['division'] = ''
            context['data'] = ''
            context['email'] = ''

            context['name1'] = ''
            context['profession1'] = ''
            context['division1'] = ''
            context['data1'] = ''
            context['email1'] = ''

            context['name2'] = ''
            context['profession2'] = ''
            context['division2'] = ''
            context['data2'] = ''
            context['email2'] = ''

            context['name3'] = ''
            context['profession3'] = ''
            context['division3'] = ''
            context['data3'] = ''
            context['email3'] = ''

            context['name4'] = ''
            context['profession4'] = ''
            context['division4'] = ''
            context['data4'] = ''
            context['email4'] = ''

        num = num + 1

    if num != 0:
        # Заполнение шаблона данными
        doc.render(context)

        # Сохранение документа
        # doc.save("ОТ2"+ f"{row[0]}"+ ".docx")
        doc.save("ОТ2.docx")
        with open('./ОТ2.docx', 'rb') as document:
            bot.send_document('5211807364', document)
        document.close()
        context['name'] = ''
        context['profession'] = ''
        print(context['profession'])
        context['division'] = ''
        context['data'] = ''
        context['email'] = ''

        context['name1'] = ''
        context['profession1'] = ''
        context['division1'] = ''
        context['data1'] = ''
        context['email1'] = ''

        context['name2'] = ''
        context['profession2'] = ''
        context['division2'] = ''
        context['data2'] = ''
        context['email2'] = ''

        context['name3'] = ''
        context['profession3'] = ''
        context['division3'] = ''
        context['data3'] = ''
        context['email3'] = ''

        context['name4'] = ''
        context['profession4'] = ''
        context['division4'] = ''
        context['data4'] = ''
        context['email4'] = ''

    # df.to_excel('./training_report.xlsx', sheet_name=message.text, index=False)
    # doc = open(r'D:\проекты питон\severalmaz_training_center_bot\training_report.xlsx',
    #                            'rb')
    # bot.send_document(message.from_user.id, doc)







def menu(id):
    bot.send_message(id, text='Вот что я могу:', reply_markup=function0)
    global key_memo
    key_memo = 1
    # bot.send_message(id,
    #                  text='Если в списке нет интересующего вопроса, вы можете связаться с нами по телефону '
    #                       '+7 (800) 300-61-22')
    # random_fact(id)




@bot.callback_query_handler(func=lambda call: True)
def query_handler(call):
    global key
    if (call.data == "employ"):
        key = "employ"
        bot.send_message(call.message.chat.id, 'Введите ФИО работника')
    elif (call.data == "memo"):
        bot.send_message(call.message.chat.id, "Укажите номер месяца за, на который необходимо сформировать документы",
                         reply_markup=test_quition)
        # memo(call.message.chat.id)
#


@bot.message_handler(content_types=['text'])
def text(message):
    global month
    if (key == "employ"):

        # Запрос на ОТ предстоящие образовательные курсы
        # SELECT employee.full_name, code_profession.profession, lerning_division.date, plane_OT.number
        #     FROM employee
        #     JOIN plane_OT ON employee.ID = plane_OT.ID
        #     JOIN code_profession ON plane_OT.profession = code_profession.ID
        #     JOIN lerning_division ON plane_OT.date = lerning_division.month
        #     WHERE employee.ID = 3


        print(message.text)
        con = sqlite3.connect('BdTrainingCenter.db')
        cursorEmployee = con.cursor()

        # запрос на человека какие курсы по ОТ проходил
        # SELECT employee.full_name, code_profession.profession, training_report_OT.date, training_report_OT.curse_num
        #     FROM employee
        #     JOIN training_report_OT ON employee.ID = training_report_OT.ID_employee
        #     JOIN code_profession ON training_report_OT.curse_profession = code_profession.ID
        #     WHERE employee.ID = 1

        bot.send_message(message.chat.id, 'Введите ФИО сотрудника по которому не обходимо составить отчет')
        sqlite_select_employee = """SELECT * from employee"""
        cursorEmployee.execute(sqlite_select_employee)
        for row_emp in cursorEmployee.fetchall():
            if row_emp[1] == message.text:
                id = row_emp[0]
                cursorReport = con.cursor()
                sqlite_select_report = """SELECT * from training_report"""
                cursorReport.execute(sqlite_select_report)

                # Открываем существующий файл
                # clean_xlsx()

                # df = pd.DataFrame({'Сотрудники': [],
                #                    'Должность организации': [],
                #                    'Реквизиты документа': [],
                #                    'Дата': []})
                df = pd.DataFrame(columns=['Сотрудники',
                                   'Должность организации',
                                   'Реквизиты документа',
                                   'Дата'])

                for row_rep in cursorReport.fetchall():
                    if row_rep[0] == id:
                        # print(row_rep)

                        cursorProfession = con.cursor()
                        sqlite_select_Profession = """SELECT * from code_profession"""
                        cursorProfession.execute(sqlite_select_Profession)
                        for row_pro in cursorProfession.fetchall():
                            if row_emp[2] == row_pro[0]:
                                # print(row_pro)
                                # df = pd.DataFrame(columns=['Имя', 'Возраст', 'Город'])

                                # Добавляем строки
                                # df = df.append({'Имя': 'Alice', 'Возраст': 25, 'Город': 'New York'}, ignore_index=True)
                                # df = pd.DataFrame({'Сотрудники': [message.text],
                                #                    'Должность организации': [row_pro[1]],
                                #                    'Реквизиты документа': [row_rep[1]],
                                #                    'Дата': row_rep[3]})
                                # df.loc= [[message.text], [row_pro[1]],[row_rep[1]],row_rep[3]]
                                # df = df.append({'Сотрудники': [message.text],
                                #                    'Должность организации': [row_pro[1]],
                                #                    'Реквизиты документа': [row_rep[1]],
                                #                    'Дата': row_rep[3]})

                                # df = df.append(pd.DataFrame({'Сотрудники': [message.text],
                                #                              'Должность организации': [row_pro[1]],
                                #                              'Реквизиты документа': [row_rep[1]],
                                #                              'Дата': [row_rep[3]]}), ignore_index=True)

                                # Создаем новый DataFrame для добавляемой строки
                                new_row = pd.DataFrame({
                                    'Сотрудники': [message.text],
                                    'Должность организации': [row_pro[1]],
                                    'Реквизиты документа': [row_rep[1]],
                                    'Дата': [row_rep[3]]
                                })

                                # Используем pd.concat для добавления новой строки
                                df = pd.concat([df, new_row], ignore_index=True)

                                # print(df)

                df.to_excel('./training_report.xlsx', sheet_name=message.text, index=False)
                doc = open(r'D:\проекты питон\severalmaz_training_center_bot\training_report.xlsx',
                           'rb')
                bot.send_document(message.from_user.id, doc)
                doc.close()

                return
    elif (message.text == "ОТ 2"):
        memoOT2(message.chat.id, month)
    elif (int(message.text) >= 1 and int(message.text) < 13):
        print(int(message.text))
        month = int(message.text)
        bot.send_message(message.chat.id, "По какому типу обучения составить записку?",
                         reply_markup=type_lerning)


if __name__ == "__main__":
    bot.polling(none_stop=True)
