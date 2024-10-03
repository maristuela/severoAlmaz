from apscheduler.schedulers.background import BackgroundScheduler
from telegram_bot_calendar import DetailedTelegramCalendar, LSTEP
from telebot import types
import telebot
import openpyxl
import pandas as pd
import time
import sqlite3
import logging
# Импортируем нужный объект из библиотеки
from docxtpl import DocxTemplate
import emoji

token = '7526701581:AAEj5YAQJ8B-_VaKAEaGjdtw4ckL_aA1u-A'
bot = telebot.TeleBot(token)
user_states = {}



# Включаем логирование
logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO)
logger = logging.getLogger(__name__)


type_new_lerning = {}


key = ""
key_memo = 0000-00-00
file_path = 'training_report.xlsx'  # Укажите путь к вашему файлу

doc = DocxTemplate("Форма СЗ на обучение по ОТ (ОТ обучение 2).docx")
doc1 = DocxTemplate("Шаблон ОТ 1.docx")
docpb = DocxTemplate("ПБшаблон.docx")
# doc2 = DocxTemplate("ПБ_шаблон.docx")

oneDate = 0
twoDate = 0
context = {
    'name': '',
    'profession': '',
    'division': '',
    'data': '',
    'email': '',
    'name1': '',
    'profession1': '',
    'division1': '',
    'data1': '',
    'email1': '',
    'name2': '',
    'profession2': '',
    'division2': '',
    'data2': '',
    'email2': '',
    'name3': '',
    'profession3': '',
    'division3': '',
    'data3': '',
    'email3': '',
    'name4': '',
    'profession4': '',
    'division4': '',
    'data4': '',
    'email4': ''
}


context1 = {
    'data': '',
    'name': '',
    'profession': '',
    'division': '',
    'newProfession': '',
    'trabl': '',
    'phone': '',
    'data1': '',
    'name1': '',
    'profession1': '',
    'division1': '',
    'newProfession1': '',
    'trabl1': '',
    'phone1': '',
    'fullName': ''

}



context2 = {
    'data': '',
    'name': '',
    'profession': '',
    'division': '',
    'document': '',
    'data1': '',
    'name1': '',
    'profession1': '',
    'division1': '',
    'document1': ''

}

# SELECT MAX(date) AS max_date FROM training_report_OT

profession = 0
month = -1
#

test_quition = telebot.types.ReplyKeyboardMarkup(True, True)
test_quition.row('1', '2', '3', '4', '5', '6')
test_quition.row('7', '8', '9', '10', '11', '12')

type_lerning = telebot.types.ReplyKeyboardMarkup(True, True)
type_lerning.row("ОТ 1", "ОТ 2", "ПБ (1-7)")

type_lerning_two = telebot.types.ReplyKeyboardMarkup(True, True)
type_lerning_two.row("ОТ", "ПБ")


# test_quition.row('7', '8', '9', '10', '11', '12')
# remove_keyboard = telebot.types.ReplyKeyboardRemove()

ruc = telebot.types.InlineKeyboardMarkup()
YZ = telebot.types.InlineKeyboardMarkup()
# работает
function1 = telebot.types.InlineKeyboardButton("Составить отчет по работнику", callback_data="employ")
# работает
function2 = telebot.types.InlineKeyboardButton("Составить отчет на год", callback_data="year")
# работает
function3 = telebot.types.InlineKeyboardButton("Добавить новое обучение", callback_data="curse")
# работает
function4 = telebot.types.InlineKeyboardButton("Сформировать служебную записку", callback_data="memo")
# function5 = telebot.types.InlineKeyboardButton("Скоректировать обучение по срокам", callback_data="time")
# работает
function6 = telebot.types.InlineKeyboardButton("Новый работник", callback_data="add_employee")

function7 = telebot.types.InlineKeyboardButton("Добавить информацию по обучению", callback_data="newLerning")
# function8 = telebot.types.InlineKeyboardButton("Добавить работника на обучение", callback_data="employeeLerning")
# работает
function9 = telebot.types.InlineKeyboardButton("Cформировать отчет по своему подразделению за период времени", callback_data="document")
ruc.add(function4).add(function9).add(function6)#.add(function8).add(function5)
YZ.add(function1).add(function2).add(function3).add(function7)


def get_professions(state):
    conn = sqlite3.connect('BdTrainingCenter.db')
    cursor = conn.cursor()
    cursor.execute(f"SELECT ID, profession FROM code_profession")
    professions = cursor.fetchall()
    conn.close()
    return professions


def add_employee(full_name, profession_id, division, email):
    conn = sqlite3.connect('BdTrainingCenter.db')
    cursor = conn.cursor()

    cursor.execute("SELECT MAX(ID) FROM employee")
    last_id = cursor.fetchone()[0]

    if last_id is None:
        last_id = 0

    new_id = last_id + 1

    cursor.execute("INSERT INTO employee (ID, full_name, profession, division, email) VALUES (?, ?, ?, ?, ?)",
                   (new_id, full_name, profession_id, division, email))

    conn.commit()
    conn.close()

    return new_id


def clean_xlsx():
    workbook = openpyxl.load_workbook(file_path)
    # Перебираем все листы в книге
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]

        # Очищаем содержимое всех ячеек
        for row in worksheet.iter_rows():
            for cell in row:
                cell.value = None  # Устанавливаем значение ячейки в None

    # Сохраняем изменения в новый файл или перезаписываем существующий
    workbook.save('training_report.xlsx')  # Укажите имя для сохранения
    # workbook.save(file_path)  # Чтобы перезаписать оригинальный файл

    print("Документ очищен.")






# Функция, отправляющая сообщение
def send_monthly_message():
    con = sqlite3.connect('BdTrainingCenter.db')
    cursor = con.cursor()
    # bot.send_message(message.chat.id, 'Введите ФИО сотрудника по которому не обходимо составить отчет')

    sqlite_select = f"""SELECT * FROM manager"""
    cursor.execute(sqlite_select)
    rows = cursor.fetchall()
    for row in rows:  # Проверяем, что строки существуют
        memoPB(row[4], 11)
        memoOT1(row[4], 11)
        memoOT2(row[4], 11)
    message = "Это ежемесячное сообщение!"
    # for chat_id in chat_ids:
    #     try:
    #         bot.send_message(chat_id, message)
    #     except Exception as e:
    #         logger.error(f"Ошибка при отправке сообщения в чат {chat_id}: {e}")



def main():
    # Устанавливаем расписление
    scheduler = BackgroundScheduler()
    # Рассылка будет происходить 1-го числа каждого месяца в 10:00
    scheduler.add_job(send_monthly_message, 'cron', day='3', hour='3', minute = '22')



    # scheduler.add_job(calender.memoPB(), 'cron', day='1', hour='10', minute='0')
    # scheduler.add_job(global_algoritmOT.update, 'cron', month='1', day='1', hour='0', minute='0')
    scheduler.start()

    # Запускаем бота
    bot.polling()

@bot.message_handler(commands=['start'])
def start_message(message):
    # con = sqlite3.connect('CityDiscoveries.db')
    # cursorObj = con.cursor()
    # albums = (message.chat.id, 0)
    # cursorObj.execute("INSERT INTO men VALUES (?,?)", albums)
    # con.commit()
    print(message.chat.id)
    print(message.chat)
    # bot.send_message(message.chat.id, 'Введите ваше ФИО')
    # bot.send_message(message.chat.id, 'Вы авторизовались, как работник ОЦ')

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    button = types.KeyboardButton("Поделиться номером", request_contact=True)
    markup.add(button)
    bot.send_message(message.chat.id, "Здравствуйте! Чтобы вы могли авторизоваться, нам нужен ваш номер. Нажмите на кнопку снизу,чтобы поделиться номером",
                     reply_markup=markup)


# Обработка получения контакта
@bot.message_handler(content_types=['contact'])
def handle_contact(message):
    global profession
    phone_number = message.contact.phone_number

    con = sqlite3.connect('BdTrainingCenter.db')
    cursor = con.cursor()
    # bot.send_message(message.chat.id, 'Введите ФИО сотрудника по которому не обходимо составить отчет')

    sqlite_select = f"""SELECT * FROM manager WHERE phone = {phone_number}"""
    cursor.execute(sqlite_select)
    rows = cursor.fetchall()
    if rows:  # Проверяем, что строки существуют
        row = rows[0]
        print(row)
        if (row[2] == -1):
            profession = -1
            bot.send_message(message.chat.id, 'Вы успешно авторизовались, как работник образовательного центра')
            print(profession)
            sqlite_select = f"""UPDATE manager SET tg = {message.chat.id} WHERE phone = {phone_number}"""
            cursor.execute(sqlite_select)
            con.commit()
            menu(message.chat.id)
            # bot.send_message(message.chat.id, 'Успешная авторизация')
        if (row[2] == 1):
            profession = 1
            bot.send_message(message.chat.id, 'Вы успешно авторизовались, как руководитель подразделения')
            print(profession)
            sqlite_select = f"""UPDATE manager SET tg = {message.chat.id} WHERE phone = {phone_number}"""
            cursor.execute(sqlite_select)
            con.commit()
            menu(message.chat.id)
            # bot.send_message(message.chat.id, 'Успешная авторизация')




    else :
        bot.send_message(message.chat.id, 'Вас нет в списке сотрудников')
    # bot.send_message(message.chat.id, f"Спасибо! Ваш номер телефона: {phone_number}")



@bot.message_handler(commands=['menu'])
def menu(message):
    global profession
    if (profession == 0):
        print(message.chat.id)
        sqlite_select = f"""SELECT * FROM manager WHERE tg = {message.chat.id}"""
        con = sqlite3.connect('BdTrainingCenter.db')
        cursor = con.cursor()
        cursor.execute(sqlite_select)
        for row in cursor.fetchall():
            if (row[2] == -1):
                profession = -1
            else:
                profession = 1

    if (profession == 1):
        menu(message.chat.id)
    if (profession == -1):
        menu(message.chat.id)
    print(profession)

def menu(id):
    print(profession)
    if (profession == -1):
        bot.send_message(id, text='Вот что я могу:', reply_markup=YZ)
    elif (profession == 1):
        bot.send_message(id, text='Вот что я могу:', reply_markup=ruc)
    global key_memo
    key_memo = 1
    # bot.send_message(id,
    #                  text='Если в списке нет интересующего вопроса, вы можете связаться с нами по телефону '
    #                       '+7 (800) 300-61-22')
    # random_fact(id)



# функция для заполнение документа типа ОТ2
def memoOT2(id, date):
    sqlite_select_tg = f"""SELECT * FROM manager WHERE tg = {id}"""
    con_tg = sqlite3.connect('BdTrainingCenter.db')
    cursor_tg = con_tg.cursor()
    cursor_tg.execute(sqlite_select_tg)
    tg = cursor_tg.fetchall()[0][2]
    print(tg)
    print(date)

    num = 0

    con = sqlite3.connect('BdTrainingCenter.db')
    cursor = con.cursor()
    if (date < 10):
        new_date = '0'+str(date)
    else:
        new_date = date
    print(new_date)
    sqlite_select = f"""SELECT employee.full_name, code_profession.Profession, employee.division, plane_OT.start_date, employee.email, plane_OT.number
FROM employee
JOIN plane_OT ON employee.ID = plane_OT.ID
JOIN code_profession ON employee.profession = code_profession.ID
WHERE employee.division = {tg} AND strftime('%m', plane_OT.start_date) = '{new_date}' AND plane_OT.number = 2
"""
    cursor.execute(sqlite_select)
    # for row_emp in cursorEmployee.fetchall():
    #     if row_emp[3] == 1:
    #         for row_emp in cursorEmployee.fetchall():

    # Данные для заполнения шаблона
    for row in cursor.fetchall():
        print(row)
        if num == 0:
            # 'name': '',
            #     'profession': '',
            #     'division': '',
            #     'data': '',
            #     'email': '',
            context['name'] = row[0]
            context['profession'] = row[1]
            print(context['profession'])
            context['division'] = row[2]
            context['data'] = row[3]
            context['email'] = row[4]
        elif num == 1:
            # 'name': '',
            #     'profession': '',
            #     'division': '',
            #     'data': '',
            #     'email': '',
            context['name1'] = row[0]
            context['profession1'] = row[1]
            context['division1'] = row[2]
            context['data1'] = row[3]
            context['email1'] = row[4]
        elif num == 2:
            # 'name': '',
            #     'profession': '',
            #     'division': '',
            #     'data': '',
            #     'email': '',
            context['name2'] = row[0]
            context['profession2'] = row[1]
            context['division2'] = row[2]
            context['data2'] = row[3]
            context['email2'] = row[4]
        elif num == 3:
            # 'name': '',
            #     'profession': '',
            #     'division': '',
            #     'data': '',
            #     'email': '',
            context['name3'] = row[0]
            context['profession3'] = row[1]
            context['division3'] = row[2]
            context['data3'] = row[3]
            context['email3'] = row[4]
        else:
            # 'name': '',
            #     'profession': '',
            #     'division': '',
            #     'data': '',
            #     'email': '',
            context['name4'] = row[0]
            context['profession4'] = row[1]
            context['division4'] = row[2]
            context['data4'] = row[3]
            context['email4'] = row[4]
            num = -1
            # Заполнение шаблона данными
            doc.render(context)

            # Сохранение документа
            doc.save("ОТ2.docx")
            with open('./ОТ2.docx', 'rb') as document:
                bot.send_document(id, document)
            document.close()
            context['name'] = ''
            context['profession'] = ''
            print(context['profession'])
            context['division'] = ''
            context['data'] = ''
            context['email'] = ''

            context['name1'] = ''
            context['profession1'] = ''
            context['division1'] = ''
            context['data1'] = ''
            context['email1'] = ''

            context['name2'] = ''
            context['profession2'] = ''
            context['division2'] = ''
            context['data2'] = ''
            context['email2'] = ''

            context['name3'] = ''
            context['profession3'] = ''
            context['division3'] = ''
            context['data3'] = ''
            context['email3'] = ''

            context['name4'] = ''
            context['profession4'] = ''
            context['division4'] = ''
            context['data4'] = ''
            context['email4'] = ''

        num = num + 1

    if num != 0:
        # Заполнение шаблона данными
        doc.render(context)

        # Сохранение документа
        # doc.save("ОТ2"+ f"{row[0]}"+ ".docx")
        doc.save("ОТ2.docx")
        with open('./ОТ2.docx', 'rb') as document:
            bot.send_document(id, document)
        document.close()
        context['name'] = ''
        context['profession'] = ''
        print(context['profession'])
        context['division'] = ''
        context['data'] = ''
        context['email'] = ''

        context['name1'] = ''
        context['profession1'] = ''
        context['division1'] = ''
        context['data1'] = ''
        context['email1'] = ''

        context['name2'] = ''
        context['profession2'] = ''
        context['division2'] = ''
        context['data2'] = ''
        context['email2'] = ''

        context['name3'] = ''
        context['profession3'] = ''
        context['division3'] = ''
        context['data3'] = ''
        context['email3'] = ''

        context['name4'] = ''
        context['profession4'] = ''
        context['division4'] = ''
        context['data4'] = ''
        context['email4'] = ''

    # df.to_excel('./training_report.xlsx', sheet_name=message.text, index=False)
    # doc = open(r'D:\проекты питон\severalmaz_training_center_bot\training_report.xlsx',
    #                            'rb')
    # bot.send_document(message.from_user.id, doc)



def memoOT1(id, date):
    sqlite_select_tg = f"""SELECT * FROM manager WHERE tg = {id}"""
    con_tg = sqlite3.connect('BdTrainingCenter.db')
    cursor_tg = con_tg.cursor()
    cursor_tg.execute(sqlite_select_tg)
    tg = cursor_tg.fetchall()[0][2]
    print(tg)
    print(date)

    num = 0

    con = sqlite3.connect('BdTrainingCenter.db')
    cursor = con.cursor()
    if (date < 10):
        new_date = '0'+str(date)
    else:
        new_date = date
    print(new_date)
    sqlite_select = f"""SELECT 
plane_OT.start_date,
employee.full_name,
employee_profession.Profession AS employee_profession,
employee.division,
plane_profession.Profession AS plane_profession,
employee.email
FROM 
employee 
JOIN 
plane_OT ON employee.ID = plane_OT.ID
JOIN 
code_profession AS employee_profession ON employee.profession = employee_profession.ID 
JOIN 
code_profession AS plane_profession ON plane_OT.profession = plane_profession.ID 
WHERE 
employee.division = {tg} 
AND strftime('%m', plane_OT.start_date) = '{new_date}' AND plane_OT.number = 1

"""
    cursor.execute(sqlite_select)
    # for row_emp in cursorEmployee.fetchall():
    #     if row_emp[3] == 1:
    #         for row_emp in cursorEmployee.fetchall():

    # Данные для заполнения шаблона
    for row in cursor.fetchall():
        print(row)
        if num == 0:
            #     'data': '',
            #     'name': '',
            #     'profession': '',
            #     'division': '',
            #     'newProfession': '',
            #     'trabl': '',
            #     'phone': '',
            context1['data'] = row[0]
            context1['name'] = row[1]
            print(context['data'])
            context1['profession'] = row[2]
            context1['division'] = row[3]
            context1['newProfession'] = row[4]
            context1['trabl'] = ' '
            context1['phone'] = row[5]
        elif num == 1:
            # 'name': '',
            #     'profession': '',
            #     'division': '',
            #     'data': '',
            #     'email': '',
            context1['data1'] = row[0]
            context1['name1'] = row[1]
            # print(context['profession'])
            context1['profession1'] = row[2]
            context1['division1'] = row[3]
            context1['newProfession1'] = row[4]
            context1['trabl1'] = ' '
            context1['phone1'] = row[5]
            sqlite_select_tg = f"""SELECT full_name FROM manager WHERE tg = {id}"""
            con_tg = sqlite3.connect('BdTrainingCenter.db')
            cursor_tg = con_tg.cursor()
            cursor_tg.execute(sqlite_select_tg)
            tg = cursor_tg.fetchall()[0][0]
            print(tg)
            print(date)
            context1['fullName'] =tg
            num = -1
            # Заполнение шаблона данными
            doc1.render(context1)

            # Сохранение документа
            # doc.save("ОТ2"+ f"{row[0]}"+ ".docx")
            doc1.save("ОТ1.docx")
            with open('./ОТ1.docx', 'rb') as document:
                bot.send_document(id, document)
            document.close()
            context1['data'] = ''
            context1['name'] = ''
            # print(context['profession'])
            context1['profession'] = ''
            context1['division'] = ''
            context1['newProfession'] = ''
            context1['trabl'] = ''
            context1['phone'] = ''
            # print(context['profession'])
            context1['profession1'] = ''
            context1['division1'] = ''
            context1['newProfession1'] = ''
            context1['trabl1'] = ''
            context1['phone1'] = ''
            context1['data1'] = ''
            context1['name1'] = ''
        num = num + 1

    if num != 0:
        sqlite_select_tg = f"""SELECT full_name FROM manager WHERE tg = {id}"""
        con_tg = sqlite3.connect('BdTrainingCenter.db')
        cursor_tg = con_tg.cursor()
        cursor_tg.execute(sqlite_select_tg)
        tg = cursor_tg.fetchall()[0][0]
        print(tg)
        print(date)
        context1['fullName'] = tg
        # Заполнение шаблона данными
        doc1.render(context1)

        # Сохранение документа
        # doc.save("ОТ2"+ f"{row[0]}"+ ".docx")
        doc1.save("ОТ1.docx")
        with open('./ОТ1.docx', 'rb') as document:
            bot.send_document(id, document)
        document.close()
        context1['name'] = ''
        context1['profession'] = ''
        print(context1['profession'])
        context1['division'] = ''
        context1['data'] = ''
        context1['email'] = ''

        context1['name1'] = ''
        context1['profession1'] = ''
        context1['division1'] = ''
        context1['data1'] = ''
        context1['email1'] = ''


    # df.to_excel('./training_report.xlsx', sheet_name=message.text, index=False)
    # doc = open(r'D:\проекты питон\severalmaz_training_center_bot\training_report.xlsx',
    #                            'rb')
    # bot.send_document(message.from_user.id, doc)



def memoPB(id, date):
    sqlite_select_tg = f"""SELECT * FROM manager WHERE tg = {id}"""
    con_tg = sqlite3.connect('BdTrainingCenter.db')
    cursor_tg = con_tg.cursor()
    cursor_tg.execute(sqlite_select_tg)
    tg = cursor_tg.fetchall()[0][2]
    print(tg)
    print(date)

    num = 0

    con = sqlite3.connect('BdTrainingCenter.db')
    cursor = con.cursor()
    if (date < 10):
        new_date = '0'+str(date)
    else:
        new_date = date
    print(new_date)
    sqlite_select = f"""SELECT plane_PB.end_date, employee.full_name, code_profession.Profession, employee.division, training_report_PB.details_document
FROM employee
JOIN plane_PB ON employee.ID = plane_PB.ID
JOIN training_report_PB ON training_report_PB.ID_employee = employee.ID
JOIN code_profession ON employee.profession = code_profession.ID
WHERE employee.division = {tg} AND strftime('%m', plane_PB.start_date) = '{new_date}'
GROUP BY employee.ID

"""
    cursor.execute(sqlite_select)
    # for row_emp in cursorEmployee.fetchall():
    #     if row_emp[3] == 1:
    #         for row_emp in cursorEmployee.fetchall():

    # Данные для заполнения шаблона
    for row in cursor.fetchall():
        print(row)
        if num == 0:
            #     'data': '',
            #     'name': '',
            #     'profession': '',
            #     'division': '',
            #     'newProfession': '',
            #     'trabl': '',
            #     'phone': '',
            context2['data'] = row[0]
            context2['name'] = row[1]
            print(context['data'])
            context2['profession'] = row[2]
            context2['division'] = row[3]
            context2['document'] = row[4]
        elif num == 1:
            # 'name': '',
            #     'profession': '',
            #     'division': '',
            #     'data': '',
            #     'email': '',
            context2['data1'] = row[0]
            context2['name1'] = row[1]
            # print(context['profession'])
            context2['profession1'] = row[2]
            context2['division1'] = row[3]
            context2['document1'] = row[4]
            num = -1
            # Заполнение шаблона данными
            docpb.render(context2)

            # Сохранение документа
            # doc.save("ОТ2"+ f"{row[0]}"+ ".docx")
            docpb.save("ПБ.docx")
            with open('./ПБ.docx', 'rb') as document:
                bot.send_document(id, document)
            document.close()

            #  'data': '',
            #     'name': '',
            #     'profession': '',
            #     'division': '',
            #     'document': '',
            #     'data1': '',
            #     'name1': '',
            #     'profession1': '',
            #     'division1': '',
            #     'document1': ''

            context2['data'] = ''
            context2['name'] = ''
            # print(context['profession'])
            context2['profession'] = ''
            context2['division'] = ''
            context2['document'] = ''
            # print(context['profession'])
            context2['profession1'] = ''
            context2['division1'] = ''
            context2['document1'] = ''
            context2['data1'] = ''
            context2['name1'] = ''
        num = num + 1

    if num != 0:
        # Заполнение шаблона данными
        docpb.render(context2)

        # Сохранение документа
        # doc.save("ОТ2"+ f"{row[0]}"+ ".docx")
        docpb.save("ПБ.docx")
        with open('./ПБ.docx', 'rb') as document:
            bot.send_document(id, document)
        document.close()
        context2['data'] = ''
        context2['name'] = ''
        # print(context['profession'])
        context2['profession'] = ''
        context2['division'] = ''
        context2['document'] = ''
        # print(context['profession'])
        context2['profession1'] = ''
        context2['division1'] = ''
        context2['document1'] = ''
        context2['data1'] = ''
        context2['name1'] = ''


@bot.message_handler(commands=['date'])
def date(m):
    calendar, step = DetailedTelegramCalendar().build()
    bot.send_message(m.chat.id,
                     f"Select {LSTEP[step]}",
                     reply_markup=calendar)


@bot.callback_query_handler(func=DetailedTelegramCalendar.func())
def cal(c):
    global oneDate, twoDate
    result, key, step = DetailedTelegramCalendar().process(c.data)
    if not result and key:
        bot.edit_message_text(f"Select {LSTEP[step]}",
                              c.message.chat.id,
                              c.message.message_id,
                              reply_markup=key)
    elif result:
        bot.edit_message_text(f"You selected {result}",
                              c.message.chat.id,
                              c.message.message_id)
        if (oneDate == 0000-00-00):
            oneDate = result
            bot.edit_message_text("Выберите дату окончания периода",
                                  c.message.chat.id,
                                  c.message.message_id)
            date(c.message)
        else:
            twoDate = result
            sqlite_select = f"""SELECT * FROM manager WHERE tg = {c.message.chat.id}"""
            con = sqlite3.connect('BdTrainingCenter.db')
            cursor = con.cursor()
            cursor.execute(sqlite_select)
            division = 0
            for row in cursor:
                division = row[2]

# !!!!!!!
#            !!!! Заполнить отчет данными о прошедших и будущих обучения за указанный периуд!!!!!!

#              Отчет по ОТ прошедших
            df = pd.DataFrame(columns=['Сотрудники',
                                        'Номер обучения ОТ',
                                       'Должность на которую проходили обучение',
                                       'Реквизиты документа',
                                       'Дата'])

            con = sqlite3.connect('BdTrainingCenter.db')

            cursorProfession = con.cursor()
            sqlite_select_Profession = f"""SELECT employee.full_name,training_report_OT.curse_num, code_profession.profession, training_report_OT.details_document, training_report_OT.date
    FROM employee
    JOIN training_report_OT ON employee.ID = training_report_OT.ID_employee 
    JOIN code_profession ON code_profession.ID = training_report_OT.curse_profession
    WHERE employee.division = {division} AND training_report_OT.date > '{oneDate}' AND training_report_OT.date < '{twoDate}'"""
            cursorProfession.execute(sqlite_select_Profession)
            for row_pro in cursorProfession.fetchall():
                new_row = pd.DataFrame({
                                'Сотрудники': row_pro[0],
                                'Номер обучения ОТ': [row_pro[1]],
                                'Должность на которую проходили обучение': [row_pro[2]],
                                'Реквизиты документа' : [row_pro[3]],
                                'Дата': [row_pro[4]]
                            }, index=[0])

                # Используем pd.concat для добавления новой строки
                df = pd.concat([df, new_row], ignore_index=True)

                # print(df)

            cursorProfession = con.cursor()
            sqlite_select_Profession = f"""SELECT employee.full_name,plane_OT.number, code_profession.profession ,plane_OT.end_date
                FROM employee
                JOIN plane_OT ON employee.ID = plane_OT.ID
                JOIN code_profession ON code_profession.ID = plane_OT.profession
                WHERE employee.division = {division} AND plane_OT.start_date > '{oneDate}' AND plane_OT.end_date < '{twoDate}'"""
            cursorProfession.execute(sqlite_select_Profession)
            for row_pro in cursorProfession.fetchall():
                new_row = pd.DataFrame({
                    'Сотрудники': row_pro[0],
                    'Номер обучения ОТ': [row_pro[1]],
                    'Должность на которую проходили обучение': [row_pro[2]],
                    'Реквизиты документа': ["Не получен"],
                    'Дата': [row_pro[3]]
                }, index=[0])

                # Используем pd.concat для добавления новой строки
                df = pd.concat([df, new_row], ignore_index=True)

                # print(df)

            df.to_excel('./Отчет_ОТ_за_период.xlsx', sheet_name=c.message.text, index=False)
            doc = open(r'D:\проекты питон\severalmaz_training_center_bot\Отчет_ОТ_за_период.xlsx',
                       'rb')
            bot.send_document(c.message.chat.id, doc)
            doc.close()
            # Формирование для ПБ

            df = pd.DataFrame(columns=['Сотрудники',
                                       'Номер обучения ПБ',
                                       'Должность',
                                       'Реквизиты документа',
                                       'Дата'])

            con = sqlite3.connect('BdTrainingCenter.db')

            cursorProfession = con.cursor()

            sqlite_select_Profession = f"""SELECT employee.full_name,training_report_PB.curse_num, code_profession.profession, training_report_PB.details_document, training_report_PB.date
            FROM employee
            JOIN training_report_PB ON employee.ID = training_report_PB.ID_employee 
            JOIN code_profession ON code_profession.ID = employee.profession
            WHERE employee.division = {division} AND training_report_PB.date > '{oneDate}' AND training_report_PB.date < '{twoDate}'"""
            cursorProfession.execute(sqlite_select_Profession)
            for row_pro in cursorProfession.fetchall():
                new_row = pd.DataFrame({
                    'Сотрудники': row_pro[0],
                    'Номер обучения ПБ': [row_pro[1]],
                    'Должность': [row_pro[2]],
                    'Реквизиты документа': [row_pro[3]],
                    'Дата': [row_pro[4]]
                }, index=[0])

                # Используем pd.concat для добавления новой строки
                df = pd.concat([df, new_row], ignore_index=True)

                # print(df)
            # SELECT employee.full_name,plane_OT.number, code_profession.profession ,plane_OT.end_date
            #                 FROM employee
            #                 JOIN plane_OT ON employee.ID = plane_OT.ID
            #                 JOIN code_profession ON code_profession.ID = plane_OT.profession
            #                 WHERE employee.division = {division} AND plane_OT.start_date > '{oneDate}' AND plane_OT.end_date < '{twoDate}'
            cursorProfession = con.cursor()
            sqlite_select_Profession = f"""SELECT employee.full_name,plane_PB.number, code_profession.profession, plane_PB.end_date
                        FROM employee
                        JOIN plane_PB ON employee.ID = plane_PB.ID
                        JOIN code_profession ON code_profession.ID = employee.profession
                        WHERE employee.division = {division} AND plane_PB.start_date > '{oneDate}' AND plane_PB.end_date < '{twoDate}'"""
            cursorProfession.execute(sqlite_select_Profession)
            for row_pro in cursorProfession.fetchall():
                new_row = pd.DataFrame({
                    'Сотрудники': row_pro[0],
                    'Номер обучения ПБ': [row_pro[1]],
                    'Должность': [row_pro[2]],
                    'Реквизиты документа': ["Не получен"],
                    'Дата': [row_pro[3]]
                }, index=[0])

                # Используем pd.concat для добавления новой строки
                df = pd.concat([df, new_row], ignore_index=True)

                # print(df)

            df.to_excel('./Отчет_ПБ_за_период.xlsx', sheet_name=c.message.text, index=False)
            doc = open(r'D:\проекты питон\severalmaz_training_center_bot\Отчет_ПБ_за_период.xlsx',
                       'rb')
            bot.send_document(c.message.chat.id, doc)
            doc.close()
            oneDate = 0000-00-00
            twoDate = 0000-00-00
            return



# Кнопки функций
@bot.callback_query_handler(func=lambda call: True)
def query_handler(call):
    global key
    if (call.data == "employ"):
        key = "employ"
        bot.send_message(call.message.chat.id, 'Введите ФИО сотрудника по которому не обходимо составить отчет')

    elif (call.data == "year"):
        print(33)
        con = sqlite3.connect('BdTrainingCenter.db')
        cursorEmployee = con.cursor()
        # bot.send_message(message.chat.id, 'Введите ФИО сотрудника по которому не обходимо составить отчет')
        sqlite_select_employee = f"""SELECT
    plane_OT.number,
    plane_OT.end_date,
    employee.division,
    COUNT(employee.ID) AS participant_count
FROM
    employee
JOIN
    plane_OT ON plane_OT.ID = employee.ID
GROUP BY
    plane_OT.number, plane_OT.end_date, employee.division"""
        cursorEmployee.execute(sqlite_select_employee)
        df = pd.DataFrame(columns=['Вид обучения',
                                   'Дата',
                                   'Подразделение',
                                   'Число работников'])
        for row_emp in cursorEmployee.fetchall():
            print(row_emp)
            # Создаем новый DataFrame для добавляемой строки
            new_row = pd.DataFrame({
                'Вид обучения': ['ОТ'+str(row_emp[0])],
                'Дата': [row_emp[1]],
                'Подразделение': [row_emp[2]],
                'Число работников': [row_emp[3]]
            }, index=[0])
            # Используем pd.concat для добавления новой строки
            df = pd.concat([df, new_row], ignore_index=True)
        print(1)
        sqlite_select_employee = f"""SELECT
            plane_PB.number,
            plane_PB.end_date,
            employee.division,
            COUNT(employee.ID) AS participant_count
        FROM
            employee
        JOIN
            plane_PB ON plane_PB.ID = employee.ID
        GROUP BY
            plane_PB.number, plane_PB.end_date, employee.division"""
        cursorEmployee.execute(sqlite_select_employee)
        for row_emp in cursorEmployee.fetchall():
            print(row_emp)
            # Создаем новый DataFrame для добавляемой строки
            new_row = pd.DataFrame({
                'Вид обучения': ['ПБ' + str(row_emp[0])],
                'Дата': [row_emp[1]],
                'Подразделение': [row_emp[2]],
                'Число работников': [row_emp[3]]
            }, index=[0])

            # Используем pd.concat для добавления новой строки
            df = pd.concat([df, new_row], ignore_index=True)
        df.to_excel('./year.xlsx', sheet_name= '1 лист', index=False)
        doc = open(r'D:\проекты питон\severalmaz_training_center_bot\year.xlsx',
                   'rb')
        bot.send_document(call.message.chat.id, doc)
        doc.close()


    elif (call.data == "curse"):
        key = "curse"
        bot.send_message(call.message.chat.id,
                         "Выберите тип обучения",
                         reply_markup=type_lerning_two)


    elif (call.data == "memo"):
        key = "memo"
        bot.send_message(call.message.chat.id, "Укажите номер месяца за, на который необходимо сформировать документы",
                         reply_markup=test_quition)

    elif (call.data == "employeeLerning"):
        key = "employeeLerning"
        bot.send_message(call.message.chat.id, 'Введите ФИО работника')
        date()
        # memo(call.message.chat.id)




    elif (call.data == "document"):
        bot.send_message(call.message.chat.id, 'Выберите дату начала периода')
        date(call.message)


    elif  (call.data == "add_employee"):
        user_states[call.message.chat.id] = {'step': 1}
        bot.reply_to(call.message, "Введите ФИО сотрудника")
    elif (key == "profession"):
        chat_id = call.message.chat.id
        prof_id = call.data.split('_')[1]
        user_states[chat_id]['profession_id'] = int(prof_id)
        bot.answer_callback_query(call.id)
        bot.send_message(chat_id, "Профессия выбрана! \nВведите email:")
        user_states[chat_id]['step'] = 3

    elif (key == "curse_4"):
        # chat_id = call.message.chat.id
        prof_id = call.data.split('_')[1]
        type_new_lerning['profession_id'] = int(prof_id)
        conn = sqlite3.connect('BdTrainingCenter.db')
        cursor = conn.cursor()
        bot.send_message(call.message.chat.id, "Обучение добавленно для одной профессии, чтобы добавить ещё профессии  нажмите +")

        cursor.execute("INSERT INTO lerning_profession (ID, type, number) VALUES (?, ?, ?)",
                       (type_new_lerning['profession_id'], type_new_lerning['type'], type_new_lerning['number']))

        conn.commit()
        conn.close()





@bot.message_handler(content_types=['text'])
def text(message):
    global month
    global key
    global type_lerning
    if (message.text == "ОТ 2"):
        memoOT2(message.chat.id, month)
    elif (message.text == "ОТ 1"):
        memoOT1(message.chat.id, month)
    elif (message.text == "ПБ (1-7)"):
        memoPB(message.chat.id, month)
    elif (key == "employ"):

        print(message.text)
        con = sqlite3.connect('BdTrainingCenter.db')
        cursorEmployee = con.cursor()
        # bot.send_message(message.chat.id, 'Введите ФИО сотрудника по которому не обходимо составить отчет')
        sqlite_select_employee = f"""SELECT employee.full_name, code_profession.profession, training_report_OT.details_document, training_report_OT.date, training_report_OT.curse_num
from employee
JOIN code_profession ON code_profession.ID = employee.profession
JOIN training_report_OT ON training_report_OT.ID_employee = employee.ID
WHERE employee.full_name = '{message.text}'"""
        cursorEmployee.execute(sqlite_select_employee)
        df = pd.DataFrame(columns=['Сотрудники',
                                   'Должность организации',
                                   'Реквизиты документа',
                                   'Дата',
                                   'Тип обучения'])
        for row_emp in cursorEmployee.fetchall():

            new_row = pd.DataFrame({
                                    'Сотрудники': [message.text],
                                    'Должность организации': [row_emp[1]],
                                    'Реквизиты документа': [row_emp[1]],
                                    'Дата': [row_emp[3]],
                                    'Тип обучения': ['ОТ' + str(row_emp[4])]
                                })

            # Используем pd.concat для добавления новой строки
            df = pd.concat([df, new_row], ignore_index=True)

                                # print(df)
        sqlite_select_employee = f"""SELECT employee.full_name, code_profession.profession, training_report_PB.details_document, training_report_PB.date, training_report_PB.curse_num
        from employee
        JOIN code_profession ON code_profession.ID = employee.profession
        JOIN training_report_PB ON training_report_PB.ID_employee = employee.ID
        WHERE employee.full_name = '{message.text}'"""
        cursorEmployee.execute(sqlite_select_employee)
        for row_emp in cursorEmployee.fetchall():
            new_row = pd.DataFrame({
                'Сотрудники': [message.text],
                'Должность организации': [row_emp[1]],
                'Реквизиты документа': [row_emp[1]],
                'Дата': [row_emp[3]],
                'Тип обучения': ['ПБ'+ str(row_emp[4])]
            })

            # Используем pd.concat для добавления новой строки
            df = pd.concat([df, new_row], ignore_index=True)

        sqlite_select_employee = f"""SELECT employee.full_name, code_profession.profession, plane_OT.end_date, plane_OT.number
        from employee
        JOIN code_profession ON code_profession.ID = employee.profession
        JOIN plane_OT ON plane_OT.ID = employee.ID
        WHERE employee.full_name = '{message.text}'"""
        cursorEmployee.execute(sqlite_select_employee)
        for row_emp in cursorEmployee.fetchall():
            new_row = pd.DataFrame({
                'Сотрудники': [message.text],
                'Должность организации': [row_emp[1]],
                'Реквизиты документа': ['Не получен'],
                'Дата': [row_emp[2]],
                'Тип обучения': ['ОТ' + str(row_emp[3])]
            })

            # Используем pd.concat для добавления новой строки
            df = pd.concat([df, new_row], ignore_index=True)

        sqlite_select_employee = f"""SELECT employee.full_name, code_profession.profession, plane_PB.end_date, plane_PB.number
                from employee
                JOIN code_profession ON code_profession.ID = employee.profession
                JOIN plane_PB ON plane_PB.ID = employee.ID
                WHERE employee.full_name = '{message.text}'"""
        cursorEmployee.execute(sqlite_select_employee)
        for row_emp in cursorEmployee.fetchall():
            new_row = pd.DataFrame({
                'Сотрудники': [message.text],
                'Должность организации': [row_emp[1]],
                'Реквизиты документа': ['Не получен'],
                'Дата': [row_emp[2]],
                'Тип обучения': ['ПБ' + str(row_emp[3])]
            })

            # Используем pd.concat для добавления новой строки
            df = pd.concat([df, new_row], ignore_index=True)

        df.to_excel('./training_report.xlsx', sheet_name=message.text, index=False)
        doc = open(r'D:\проекты питон\severalmaz_training_center_bot\training_report.xlsx',
                           'rb')
        bot.send_document(message.from_user.id, doc)
        doc.close()

        return
    elif key == "memo":
        if (message.text.isdigit()):
            if (int(message.text) >= 1 and int(message.text) < 13):
                # if (time.time())
                bot.send_message(message.chat.id, "По какому типу обучения составить записку?",
                         reply_markup=type_lerning)
                month = int(message.text)
    elif (key == "curse"):
        type_new_lerning['type'] = str(message.text)
        key = "curse_1"
        bot.send_message(message.chat.id,
                                 "Выведите номер обучения")
    elif (key == "curse_1"):
        type_new_lerning['number'] = message.text
        key = "curse_2"
        bot.send_message(message.chat.id, "Выведите переодичность обучения в годах")
    elif (message.text == "+"):
        key = "curse_4"
        conn = sqlite3.connect('BdTrainingCenter.db')
        cursor = conn.cursor()
        cursor.execute(f"SELECT ID, profession FROM code_profession")
        professions = cursor.fetchall()
        conn.close()
        markup = types.InlineKeyboardMarkup()

        for prof_id, prof_name in professions:
            markup.add(types.InlineKeyboardButton(prof_name, callback_data=f'profession_{prof_id}'))
        bot.send_message(message.chat.id, "Выбирите профессии, которые должны обучаться новому обучению",
                         reply_markup=markup)

    elif (key == "curse_2"):
        print(3545645)
        type_new_lerning['year'] = int(message.text)
        conn = sqlite3.connect('BdTrainingCenter.db')
        cursor = conn.cursor()
        bot.send_message(message.chat.id, "Добавленно новое обучение")

        cursor.execute("INSERT INTO curse (type, number, year) VALUES (?, ?, ?)",
        (type_new_lerning['type'], type_new_lerning['number'], type_new_lerning['year']))

        conn.commit()
        conn.close()
        key = "curse_4"
        conn = sqlite3.connect('BdTrainingCenter.db')
        cursor = conn.cursor()
        cursor.execute(f"SELECT ID, profession FROM code_profession")
        professions = cursor.fetchall()
        conn.close()
        markup = types.InlineKeyboardMarkup()

        for prof_id, prof_name in professions:
            markup.add(types.InlineKeyboardButton(prof_name, callback_data=f'profession_{prof_id}'))
        bot.send_message(message.chat.id, "Выбирите профессии, которые должны обучаться новому обучению", reply_markup=markup)
            # # отчет по человеку
    state = {'step': 0}
    try:
        state = user_states[message.chat.id]
    except:
        print(3454)

    if state['step'] == 1:
        state['full_name'] = message.text

        # Fetch professions and create buttons
        bot.send_message(message.chat.id, "Введите номер подразделение")
        state['step'] = 2
        # key = "new_division"
    elif state['step'] == 2:
        state['division'] = message.text
        # state['step'] = 2
        professions = get_professions(state)
        markup = types.InlineKeyboardMarkup()

        for prof_id, prof_name in professions:
            markup.add(types.InlineKeyboardButton(prof_name, callback_data=f'profession_{prof_id}'))

        bot.send_message(message.chat.id, "Выберите профессию сотрудника", reply_markup=markup)
        key = "profession"

    elif state['step'] == 3:
        # state['division'] = 1
        state['email'] = message.text
        new_id = add_employee(state['full_name'], state['profession_id'], state['division'], state['email'])

        bot.reply_to(message, f"Новый сотрудник добавлен с ID: {new_id}")
        del user_states[message.chat.id]

    elif (key == "employeeLerning"):
        con1 = sqlite3.connect('BdTrainingCenter.db')
        cursor1 = con1.cursor()
        sqlite_select = f"""SELECT ID FROM employee WHERE full_name = {message.text} """
        cursor1.execute(sqlite_select)
        con1.commit()
        if len(cursor1.fetchall()) == 0:
            bot.send_message(message.chat.id, "Работника с данным ФИО нет, проверти правописание и повторите попытку")
        else:
            bot.send_message(message.chat.id, "Записать на какой тип обучения",
                             reply_markup=type_lerning_two)


    # elif (key == "document"):
    #     if (message.text.isdigit()):
    #         bot.send_message(message.chat.id, 'Выберите дату начала периуда')
    #         date(message)







if __name__ == "__main__":
    bot.polling(none_stop=True)